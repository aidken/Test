#! /usr/bin/env python3
# -*- coding: utf-8 -*-

import sys, io
import logging
import datetime
import json
from openpyxl import load_workbook
from dataclasses import dataclass, field

# cSpell:ignore openpyxl datefmt

# consider to import mj.util helper functions
# instead of defining them locally
# from mj.util import _convert_to_date, _convert_to_str, _convert_to_int, _convert_to_float
def _convert_to_date(tmp) -> datetime.date:
    if tmp is None or tmp in ('None', ''):
        return ''
    elif isinstance(tmp, (datetime.date, datetime.datetime)):
        return tmp.date()
    elif isinstance(tmp, str):
        x = datetime.datetime.strptime(tmp, '%Y-%m-%d')
        return x.date()
    else:
        raise ValueError


def _convert_to_str(tmp) -> str:
    if tmp is None or tmp in ('None', ''):
        return ''
    else:
        return str(tmp).strip()


def _convert_to_int(tmp) -> int:
    if tmp is None or tmp in ('None', ''):
        return 0
    else:
        return int(tmp)


def _convert_to_float(tmp) -> float:
    if tmp is None or tmp in ('None', ''):
        return 0
    else:
        return float(tmp)


class Record_encoder(json.JSONEncoder):
    # JSON Encoder
    def default(self, obj):
        if isinstance(obj, Record):
            if isinstance(obj.some_date, datetime.date):
                some_date = obj.some_date.isoformat()
            elif obj.some_date is None or obj.some_date in ('None', ''):
                some_date = ''
            else:
                raise ValueError
            return {
                'row_number': obj.row_number,
                'some_str'  : obj.some_str,
                'some_int'  : obj.some_int,
                'some_float': obj.some_float,
                'some_date' : some_date,
            }
        else:
            return super().default(obj)


def Record_decoder(tmp):
    # JSON decoder
    return Record(
        row_number = tmp['row_number'],
        some_str   = tmp['some_str'],
        some_int   = tmp['some_int'],
        some_float = tmp['some_float'],
        some_date  = tmp['some_date'],
    )


@dataclass
class Report:
    file   : str
    records: list = field(default_factory=list, init=False)

    def __post_init__(self):
        self.file = _convert_to_str(self.file)


@dataclass
class Record:
    row_number: int
    some_str  : str
    some_int  : int
    some_float: float
    some_date : datetime.date

    def __post_init__(self):
        self.row_number = _convert_to_int(self.row_number)
        self.some_str   = _convert_to_str(self.some_str)
        self.some_int   = _convert_to_int(self.some_int)
        self.some_float = _convert_to_float(self.some_float)
        self.some_date  = _convert_to_date(self.some_date)


def parse(file, worksheet, row_of_label, callback=None):

    # iterate through rows of an Excel spreadsheet

    if callback is not None and not callable(callback):
        raise ValueError(
            f'callback given but it it not callable. It is a {type(callback)}.'
        )

    wb = load_workbook(filename=file, read_only=True, data_only=True)
    ws = wb[worksheet]
    logging.debug(f'Slurp item information. Looking at worksheet "{ws.title}".')

    r = Report(file=file)

    logging.debug(f'Max row of this sheet is {ws.max_row}.')
    for row_number, row in enumerate(ws.iter_rows(min_row=row_of_label+1, max_row=ws.max_row+1, values_only=True), start=row_of_label+1):
        # In openpyxl, row and column are one based, not zero based. Cell "A1" is row 1, col 1.
        # Note that when you iterate through rows, now a row is a tuple, and it is zero based.
        # In this example, row[0] is column A.
        x = Record(
            row_number = row_number,
            some_str   = row[0],
            some_int   = row[1],
            some_float = row[2],
            some_date  = row[3],
        )

        if callback is not None and callable(callback):
            callback(x)
        else:
            r.records.append(x)

    if callback is None:
        return r


def main():
    pass


if __name__ == '__main__':

    # https://qiita.com/jack-low/items/91bf9b5342965352cbeb
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    # logger setup
    filename = str(sys.argv[0])[:-3] + '.log'
    format   = '%(asctime)s - %(filename)s: %(lineno)s: %(funcName)s - %(levelname)-8s: %(message)s'
    logging.basicConfig(
        filename = filename,
        format   = format,
        datefmt  = '%m-%d %H:%M',
        level    = logging.INFO,
        # level    = logging.DEBUG,
        # level    = logging.ERROR,
    )

    # https://docs.python.org/3/howto/logging-cookbook.html#logging-to-multiple-destinations
    # define a Handler which writes INFO messages or higher to the sys.stderr
    console = logging.StreamHandler()
    console.setLevel(logging.INFO)
    # set a format which is simpler for console use
    formatter = logging.Formatter('%(name)-12s: %(levelname)-8s %(message)s')
    # tell the handler to use this format
    console.setFormatter(formatter)
    # add the handler to the root logger
    logging.getLogger('').addHandler(console)

    main()
